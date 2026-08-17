"""Export Codex runner log records to a standards-compliant Excel workbook."""

from __future__ import annotations

import argparse
import io
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADERS = [
    "Timestamp", "Status", "Success", "Task", "Project", "Duration (ms)",
    "Codex active (ms)", "Approval wait (ms)", "Approvals", "Policy",
    "Commands", "Task ID", "Thread ID", "Error", "Output",
]

WIDTHS = [22, 14, 10, 42, 22, 16, 18, 19, 12, 16, 12, 38, 38, 35, 60]


def safe_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = "".join(
        character
        for character in text
        if ord(character) in (9, 10, 13)
        or 0x20 <= ord(character) <= 0xD7FF
        or 0xE000 <= ord(character) <= 0xFFFD
        or 0x10000 <= ord(character) <= 0x10FFFF
    )
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def excel_timestamp(value: object) -> datetime | str:
    if not value:
        return ""
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo:
            parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
        return parsed
    except ValueError:
        return safe_text(value)


def create_workbook(logs: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Runner Logs"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"

    sheet.append(HEADERS)
    for log in logs:
        commands = log.get("commands") or []
        sheet.append([
            excel_timestamp(log.get("timestamp")),
            safe_text(log.get("status")),
            "Yes" if log.get("success") else "No",
            safe_text(log.get("task")),
            safe_text(log.get("project")),
            int(log.get("durationMs") or 0),
            int(log.get("codexActiveMs") or 0),
            int(log.get("approvalWaitMs") or 0),
            int(log.get("approvalCount") or 0),
            safe_text(log.get("approvalPolicy")),
            len(commands),
            safe_text(log.get("taskId")),
            safe_text(log.get("threadId")),
            safe_text(log.get("error")),
            safe_text(log.get("output")),
        ])

    header_fill = PatternFill("solid", fgColor="246BFD")
    header_font = Font(name="Aptos Display", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=10, color="17202B")
    divider = Side(style="thin", color="E6E9EE")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=divider)
        row[0].number_format = "yyyy-mm-dd hh:mm:ss"
        for column in (5, 6, 7, 8, 10):
            row[column].number_format = "#,##0"
            row[column].alignment = Alignment(horizontal="right", vertical="top")
        status = str(row[1].value).lower()
        row[1].fill = PatternFill("solid", fgColor="EAF8F1" if status == "completed" else "FCECEE")
        row[1].font = Font(name="Aptos", size=10, bold=True, color="208354" if status == "completed" else "BC4450")
        sheet.row_dimensions[row[0].row].height = 54

    for index, width in enumerate(WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    if logs:
        table = Table(displayName="RunnerLogsTable", ref=f"A1:O{len(logs) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    sheet.auto_filter.ref = f"A1:O{max(len(logs) + 1, 1)}"
    sheet.auto_filter.add_sort_condition(f"A2:A{max(len(logs) + 1, 2)}", descending=True)
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:1"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Codex runner logs to Excel.")
    parser.add_argument("--input", help="JSON input file; defaults to stdin")
    parser.add_argument("--output", help="XLSX output file; defaults to stdout")
    args = parser.parse_args()

    if args.input:
        with open(args.input, "r", encoding="utf-8") as source_file:
            payload = json.load(source_file)
    else:
        payload = json.load(sys.stdin)

    logs = payload.get("logs", payload) if isinstance(payload, dict) else payload
    if not isinstance(logs, list):
        raise ValueError("Expected a JSON array of log records.")

    data = create_workbook(logs)
    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(data)
    else:
        sys.stdout.buffer.write(data)


if __name__ == "__main__":
    main()
